"""
Import Excel payout data into MySQL database
Run this AFTER setup_database.py

Usage: python import_excel.py reference/Sample_Payout_Grid_Template.xlsx

Excel Format:
  • 19 columns (A-S): Company, State, RTO_Code, Vehicle_Category, Vehicle_Type, 
    Fuel_Type, Vehicle_Age, Policy_Type, Business_Type, CC_Slab, GVW_Slab, 
    Watt_Slab, Seating_Capacity, Claim_Status, NCB_Slab, CPA_Cover, 
    Zero_Depreciation, Trailer, Payout_Percentage
  • Extra columns beyond S are ignored (Payin, Gross, TDS, etc.)
  • Data starts from Row 2 (Row 1 is headers)
  • One sheet with data per state (or combine all into one)
  • INSTRUCTIONS sheet is skipped

Comma-Separated Values (Cartesian Product):
  • ANY column can contain comma-separated values
  • System automatically creates all combinations
  • Example: RTO "1,2" + Vehicle_Type "Bike,Scooter" = 4 records
  • Time-efficient: Define criteria once with variations inline
"""

import openpyxl
from openpyxl import load_workbook
import mysql.connector
from mysql.connector import Error
import re
from config import DB_HOST, DB_PORT, DB_USER, DB_PASSWORD, DB_NAME
import sys
from itertools import product

# State to RTO mapping for "Entire State" expansion
STATE_RTOS = {
    'TN': [str(i).zfill(2) for i in range(1, 100)] + ['15M', '83M'],  # Tamil Nadu
    'PY': ['01', '02', '03', '04', '05'],  # Pondicherry
    'KA': [str(i).zfill(2) for i in range(1, 72)],  # Karnataka
    'KL': [str(i).zfill(2) for i in range(1, 74)],  # Kerala
    'AP': [str(i).zfill(2) for i in range(1, 38)],  # Andhra Pradesh
    'MH': [str(i).zfill(2) for i in range(1, 57)],  # Maharashtra
    'TS': [str(i).zfill(2) for i in range(1, 17)],  # Telangana
}

# Chennai RTOs (for "ro-chn" expansion)
CHENNAI_RTOS = {'01', '02', '03', '04', '05', '06', '07', '09', '10', '11', '12', '13', '14', 
                '18', '22', '85'}

def expand_rto_code(rto_code_str, state):
    """
    Expand RTO code string to list of individual RTOs
    
    Handles:
    - "Entire Tamilnadu" → all state RTOs
    - "ro-chn" → Rest of Chennai (all TN except explicit Chennai RTOs)
    - "01,02,03" → splits into individual codes
    - Single code "19" → single RTO
    
    Returns: list of RTOs, or empty list if invalid
    """
    if not rto_code_str or not state:
        return [rto_code_str] if rto_code_str else []
    
    rto_str = str(rto_code_str).strip()
    rto_lower = rto_str.lower()
    
    # Handle "Entire State" references
    if rto_lower.startswith('entire'):
        # "Entire Tamilnadu" or "Entire Pondicherry" etc
        if 'tamilnadu' in rto_lower or 'tamil' in rto_lower:
            return sorted(STATE_RTOS.get('TN', []))
        elif 'pondicherry' in rto_lower or 'puducherry' in rto_lower:
            return sorted(STATE_RTOS.get('PY', []))
        elif 'karnataka' in rto_lower:
            return sorted(STATE_RTOS.get('KA', []))
        elif 'kerala' in rto_lower:
            return sorted(STATE_RTOS.get('KL', []))
        else:
            # Try to use state code
            state_upper = str(state).upper().strip()
            if state_upper in STATE_RTOS:
                return sorted(STATE_RTOS[state_upper])
    
    # Handle special code "ro-chn" (Rest of Chennai)
    if rto_lower == 'ro-chn':
        return ['ro-chn']  # Keep as-is for database queries
    
    # Handle comma-separated or single RTOs
    if ',' in rto_str:
        # Split comma-separated codes, remove any trailing text
        codes = [r.strip() for r in rto_str.split(',') if r.strip()]
        cleaned = []
        for code in codes:
            if '(' in code:
                code = code.split('(')[0].strip()
            if code:
                cleaned.append(code)
        return cleaned
    
    # Single RTO code
    if '(' in rto_str:
        rto_str = rto_str.split('(')[0].strip()
    
    return [rto_str] if rto_str else []

def get_db_connection():
    """Get database connection"""
    try:
        conn = mysql.connector.connect(
            host=DB_HOST,
            port=DB_PORT,
            user=DB_USER,
            password=DB_PASSWORD,
            database=DB_NAME
        )
        return conn
    except Error as e:
        print(f"[ERROR] Database connection failed: {e}")
        return None

def get_or_create_company(cursor, company_name):
    """Get company_id or create if not exists"""
    try:
        cursor.execute("SELECT company_id FROM companies WHERE company_name = %s", (company_name,))
        result = cursor.fetchone()
        
        if result:
            return result[0]
        else:
            cursor.execute("INSERT INTO companies (company_name) VALUES (%s)", (company_name,))
            return cursor.lastrowid
    except Error as e:
        print(f"[ERROR] Error managing company: {e}")
        return None

def normalize_value(value):
    """Normalize cell values"""
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() if value.strip() else None
    return value

def normalize_fuel_type(fuel_str):
    """
    Normalize fuel type string
    Map: EV → Electric
    Keep others as-is (including spelling mistakes like "Pertol")
    """
    if not fuel_str:
        return fuel_str
    
    fuel_lower = str(fuel_str).strip().lower()
    
    # Map EV to Electric
    if fuel_lower == 'ev':
        return 'Electric'
    
    # Keep all others as-is
    return str(fuel_str).strip()

def normalize_vehicle_category_type(category, vtype):
    """
    Keep vehicle_category and vehicle_type as-is.
    Flatbed remains a vehicle_type under GCV category.
    """
    category = str(category).strip() if category else ""
    vtype = str(vtype).strip() if vtype else ""
    return (category, vtype)

def parse_excel_payouts(filepath):
    """Parse Excel file and return payout records
    
    Dynamically reads columns from Excel file header row.
    Expected columns include: Company, State, RTO_Code, Vehicle_Category, Vehicle_Type,
    Fuel_Type, Vehicle_Age, Policy_Type, Business_Type, CC_Slab, GVW_Slab, Watt_Slab,
    Seating_Capacity, Conditions, NCB_Slab, CPA_Cover, Zero_Depreciation, Trailer,
    Final Payout, and others (ignored).
    
    Special cases:
    - "Declined-XX,YY,ZZ" in RTO_Code column: RTOs where insurance is NOT available
    - Comma-separated values in ANY column: Auto-expands (Cartesian product)
    - "Conditions" column: Stores special conditions for each commission rate
    """
    records = []
    declined_rtos = []
    declined_values = []
    
    try:
        # Load with data_only=True to get calculated values instead of formulas
        wb = load_workbook(filepath, data_only=True)
        
        # Process all sheets except INSTRUCTIONS
        for sheet_name in wb.sheetnames:
            if sheet_name.upper() == "INSTRUCTIONS":
                continue
            
            ws = wb[sheet_name]
            print(f"\n[SHEET] Processing sheet: {sheet_name}")
            
            # Read header row (row 1) - dynamically determine columns
            header_row = []
            for col_idx in range(1, ws.max_column + 1):
                header_cell = ws.cell(1, col_idx).value
                header_row.append(normalize_value(header_cell))
            
            # Create column index mapping
            col_map = {name.lower(): idx for idx, name in enumerate(header_row) if name}
            
            print(f"   Columns found: {[h for h in header_row if h]}")
            
            # Verify required columns exist
            required_cols = ['company', 'state', 'rto_code', 'vehicle_type', 'fuel_type', 
                            'policy_type', 'business_type', 'ncb_slab', 'cpa_cover', 'zero_depreciation']
            missing = [col for col in required_cols if col not in col_map]
            if missing:
                    print(f"   [WARN] Missing columns: {missing}")
            
            # Read data rows (starting from row 2)
            row_count = 0
            for row_idx in range(2, ws.max_row + 1):
                row_data = {}
                
                # Extract all columns from row
                for col_name, col_idx in col_map.items():
                    cell_value = ws.cell(row_idx, col_idx + 1).value
                    row_data[col_name] = normalize_value(cell_value)

                # Generalized "Declined-" handling for ANY column
                # Accept Declined rows even if 'company' cell is empty — user may only provide declined entries.
                def normalize_declined_rto_code(code):
                    if not code:
                        return code
                    c = str(code).strip()
                    digits = ''.join([ch for ch in c if ch.isdigit()])
                    if digits:
                        return digits.zfill(2)
                    return c

                found_decline = False
                for field_name, field_val in row_data.items():
                    if not field_val or not isinstance(field_val, str):
                        continue

                    m = re.match(r'^\s*declined\s*-?\s*(.+)$', field_val, re.IGNORECASE)
                    if not m:
                        continue

                    declined_part = m.group(1).strip()
                    state = row_data.get("state", "")

                    # Split declined list ONLY on commas (ignore other separators)
                    declined_list = [item.strip() for item in declined_part.split(',') if item.strip()]
                    for item in declined_list:
                        if field_name == 'rto_code':
                            normalized = normalize_declined_rto_code(item)
                            declined_rtos.append({
                                'state': state,
                                'rto_code': normalized,
                                'reason': f'Declined - {row_data.get("company", "N/A")}'
                            })
                        else:
                            # Store generic declined parameter (field name + value)
                            declined_values.append({
                                'field_name': field_name,
                                'field_value': item,
                                'state': state,
                                'rto_code': None,
                                'reason': f'Declined - {row_data.get("company", "N/A")}'
                            })

                    found_decline = True

                if found_decline:
                    continue  # Skip further processing for this row

                # Skip empty rows (non-declined)
                if not row_data.get("company"):
                    continue
                
                # Required fields validation (relaxed: require at least company, state, rto_code)
                required_fields = ["company", "state", "rto_code"]
                
                if not all(row_data.get(field) for field in required_fields):
                    continue  # Skip rows missing the absolute minimum
                
                # Build record(s) using Cartesian product of comma-separated values
                try:
                    # Create lists of values for each field
                    field_values = {}
                    
                    # Special handling for RTO_Code - expand "Entire State" or comma-separated
                    rto_code_str = row_data.get("rto_code", "")
                    if rto_code_str:
                        state = row_data.get("state", "")
                        expanded_rtos = expand_rto_code(rto_code_str, state)
                        field_values["rto_code"] = expanded_rtos if expanded_rtos else [rto_code_str]
                        
                        if expanded_rtos and len(expanded_rtos) > 1:
                            print(f"   [OK] Expanding RTO_Code '{rto_code_str}' -> {len(expanded_rtos)} RTOs")
                    else:
                        field_values["rto_code"] = [rto_code_str]
                    
                    # Handle other fields normally (comma-separated split)
                    for field_name, field_value in row_data.items():
                        if field_name == "rto_code":  # Already handled above
                            continue
                        
                        if field_value is None:
                            field_values[field_name] = [None]
                        elif isinstance(field_value, str) and field_name not in ["payout_percentage", "final payout"]:
                            # Split comma-separated values (skip payout percentage)
                            values = [v.strip() for v in field_value.split(",") if v.strip()]
                            
                            # Apply fuel type normalization (EV → Electric)
                            if field_name == "fuel_type":
                                values = [normalize_fuel_type(v) for v in values]
                            
                            field_values[field_name] = values if values else [None]
                        else:
                            field_values[field_name] = [field_value]
                    
                    # Get payout column (could be "Final Payout" or "payout_percentage")
                    payout_col = None
                    if 'final payout' in col_map:
                        payout_col = 'final payout'
                    elif 'payout_percentage' in col_map:
                        payout_col = 'payout_percentage'
                    
                    if not payout_col:
                        print(f"   [WARN] No payout column found in row {row_idx}")
                        continue
                    
                    # Generate all combinations using Cartesian product
                    field_names = list(row_data.keys())
                    value_lists = [field_values.get(name, [None]) for name in field_names]
                    
                    # Create one record per combination
                    for combo in product(*value_lists):
                        combo_dict = dict(zip(field_names, combo))
                        
                        try:
                            # Extract conditions (normalize "No" to None)
                            conditions = str(combo_dict.get("conditions", "")).strip() if combo_dict.get("conditions") else None
                            if conditions and conditions.lower() == "no":
                                conditions = None
                            
                            # Normalize vehicle category and type (e.g., Flatbed as category)
                            normalized_cat, normalized_type = normalize_vehicle_category_type(
                                combo_dict.get("vehicle_category", ""),
                                combo_dict.get("vehicle_type", "")
                            )
                            
                            # Allow empty payout cells: treat missing/blank as 0.00
                            payout_raw = combo_dict.get(payout_col)
                            if payout_raw is None or (isinstance(payout_raw, str) and not payout_raw.strip()):
                                payout_percentage = 0.0
                            else:
                                payout_percentage = float(str(payout_raw).replace("%", "").strip())

                            # Helper to convert N/A or empty values to None for dates
                            def parse_date(val):
                                if val is None:
                                    return None
                                val_str = str(val).strip().upper()
                                if val_str in ('N/A', 'NA', '', 'NONE', '-'):
                                    return None
                                # If it's already a date object, return as-is
                                from datetime import date, datetime
                                if isinstance(val, (date, datetime)):
                                    return val
                                return val
                            
                            # Helper to clean string values (N/A -> None)
                            def clean_str(val, default=None):
                                if val is None:
                                    return default
                                val_str = str(val).strip()
                                if val_str.upper() in ('N/A', 'NA', '', 'NONE', '-'):
                                    return default
                                return val_str
                            
                            record = {
                                'company_name': str(combo_dict["company"]).strip(),
                                'state': str(combo_dict["state"]).strip(),
                                'rto_code': str(combo_dict["rto_code"]).strip()[:10],
                                'vehicle_category': normalized_cat or "General",
                                'vehicle_type': normalized_type,
                                'fuel_type': str(combo_dict["fuel_type"]).strip(),
                                'vehicle_age': clean_str(combo_dict.get("vehicle_age"), "Various"),
                                'policy_type': str(combo_dict["policy_type"]).strip(),
                                'business_type': str(combo_dict["business_type"]).strip(),
                                'cc_slab': clean_str(combo_dict.get("cc_slab")),
                                'gvw_slab': clean_str(combo_dict.get("gvw_slab")),
                                'watt_slab': clean_str(combo_dict.get("watt_slab")),
                                'seating_capacity': clean_str(combo_dict.get("seating_capacity")),
                                'ncb_slab': str(combo_dict["ncb_slab"]).strip(),
                                'cpa_cover': str(combo_dict["cpa_cover"]).strip(),
                                'zero_depreciation': str(combo_dict["zero_depreciation"]).strip(),
                                'trailer': clean_str(combo_dict.get("trailer")),
                                'conditions': conditions,
                                'payout_percentage': payout_percentage,
                                'make_model': clean_str(combo_dict.get("make")),
                                'effective_from': parse_date(combo_dict.get("date_from")),
                                'effective_to': parse_date(combo_dict.get("date_till"))
                            }
                            records.append(record)
                            row_count += 1
                        
                        except (ValueError, AttributeError) as e:
                            # Skip invalid combinations
                            continue
                
                except Exception as e:
                    print(f"   [WARN] Skipped row {row_idx}: {e}")
                    continue
            
            print(f"   [OK] Loaded {row_count} records from {sheet_name}")
        
        print(f"\n[OK] Total parsed: {len(records)} payout records")
        print(f"[INFO] Total declined RTOs: {len(declined_rtos)}")
        print(f"[INFO] Total declined generic values: {len(declined_values)}")
        return records, declined_rtos, declined_values
    
    except Exception as e:
        print(f"[ERROR] Error parsing Excel: {e}")
        return [], [], []

def normalized_value(value):
    """Normalize value for comparison"""
    if not value:
        return ""
    return str(value).strip().lower()

def import_payouts_to_db(records, declined_rtos, declined_values=None):
    """Import payout records to database with all 18 parameters

    If `records` is empty we still import declined lists (declined_rtos / declined_values).
    """
    conn = get_db_connection()
    if not conn:
        return False
    
    cursor = conn.cursor()
    inserted = 0
    skipped = 0
    
    try:
        # Insert payout records if present
        if records:
            for idx, record in enumerate(records, 1):
                try:
                    # Get or create company
                    company_id = get_or_create_company(cursor, record['company_name'])
                    if not company_id:
                        skipped += 1
                        continue
                    
                    # Insert payout record with all parameters
                    query = """
                    INSERT INTO payout_rules (
                        company_id, state, rto_code, vehicle_category, vehicle_type,
                        fuel_type, vehicle_age, policy_type, business_type, cc_slab,
                        gvw_slab, watt_slab, seating_capacity, ncb_slab,
                        cpa_cover, zero_depreciation, trailer, payout_percentage,
                        conditions, make_model, effective_from, effective_to
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """
                    
                    cursor.execute(query, (
                        company_id,
                        record['state'],
                        record['rto_code'],
                        record['vehicle_category'],
                        record['vehicle_type'],
                        record['fuel_type'],
                        record['vehicle_age'],
                        record['policy_type'],
                        record['business_type'],
                        record['cc_slab'],
                        record['gvw_slab'],
                        record['watt_slab'],
                        record['seating_capacity'],
                        record['ncb_slab'],
                        record['cpa_cover'],
                        record['zero_depreciation'],
                        record['trailer'],
                        record['payout_percentage'],
                        record.get('conditions', None),
                        record.get('make_model', None),
                        record.get('effective_from', None),
                        record.get('effective_to', None)
                    ))
                    
                    inserted += 1
                    
                    # Commit every 100 records
                    if idx % 100 == 0:
                        conn.commit()
                        print(f"   [OK] Imported {idx} records...")
                
                except Error as e:
                    skipped += 1
                    if skipped <= 5:  # Show first 5 errors
                        print(f"   [WARN] Error at record {idx}: {e}")
        else:
            print("[INFO] No payout records to import - will import declined entries only")
        
        # Final commit for any inserted payout rows
        conn.commit()
        
        # Import declined RTOs
        if declined_rtos:
            try:
                cursor = conn.cursor()
                for declined in declined_rtos:
                    query = """
                    INSERT INTO declined_rtos (state, rto_code, reason)
                    VALUES (%s, %s, %s)
                    ON DUPLICATE KEY UPDATE reason = VALUES(reason)
                    """
                    cursor.execute(query, (declined['state'], declined['rto_code'], declined['reason']))
                conn.commit()
                print(f"   [OK] Imported {len(declined_rtos)} declined RTOs")
            except Error as e:
                print(f"   [WARN] Error importing declined RTOs: {e}")

        # Import generic declined values (field-based)
        if declined_values:
            try:
                cursor = conn.cursor()
                for d in declined_values:
                    query = """
                    INSERT INTO declined_values (field_name, field_value, state, rto_code, reason)
                    VALUES (%s, %s, %s, %s, %s)
                    ON DUPLICATE KEY UPDATE reason = VALUES(reason)
                    """
                    cursor.execute(query, (d['field_name'], d['field_value'], d.get('state'), d.get('rto_code'), d.get('reason')))
                conn.commit()
                print(f"   [OK] Imported {len(declined_values)} declined values")
            except Error as e:
                print(f"   [WARN] Error importing declined values: {e}")
        
        cursor.close()
        conn.close()
        
        print(f"\n[OK] Import completed!")
        print(f"   Inserted: {inserted}")
        print(f"   Skipped: {skipped}")
        
        return True
    
    except Error as e:
        print(f"[ERROR] Import failed: {e}")
        conn.rollback()
        cursor.close()
        conn.close()
        return False

def main():
    """Main import routine"""
    print("=" * 70)
    print(" POSP PAYOUT CHECKER - IMPORT EXCEL DATA TO DATABASE")
    print("=" * 70)
    
    # Check if file path provided
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
    else:
        filepath = "Sample_Payout_Grid_Template.xlsx"
    
    print(f"\n[FILE] Source Excel: {filepath}")
    print(f"[DB] Target Database: {DB_NAME}")
    print(f"[TABLE] Target Table: payout_rules (18 parameters + payout %)\n")
    
    # Verify Excel exists
    from pathlib import Path
    if not Path(filepath).exists():
        print(f"[ERROR] File not found: {filepath}")
        print(f"   Expected location: {Path.cwd() / filepath}")
        print(f"\n[TIP] Usage: python import_excel.py <path_to_excel>")
        return
    
    # Parse Excel
    print("[PARSE] Parsing Excel file...")
    print("   Reading columns A-S (19 columns)")
    print("   Ignoring any extra columns beyond S\n")
    records, declined_rtos, declined_values = parse_excel_payouts(filepath)
    
    if not records and not declined_rtos:
        print("[ERROR] No valid records found in Excel")
        return
    
    # Show sample record
    if records:
        print(f"\n[SAMPLE] Sample record (first row):")
        sample = records[0]
        for key, value in sample.items():
            print(f"   {key:.<25} {value}")
    
    # Import to database
    print(f"\n[IMPORT] Importing {len(records)} records to database...")
    if import_payouts_to_db(records, declined_rtos, declined_values):
        print("\n" + "=" * 70)
        print("[SUCCESS] IMPORT COMPLETED SUCCESSFULLY!")
        print("\n[NEXT STEPS]:")
        print("   1. Verify data: SELECT COUNT(*) FROM payout_rules;")
        print("   2. Test API: python -m uvicorn app:app --reload")
        print("   3. Open browser: http://localhost:8000/")
        print("=" * 70)
    else:
        print("\n[ERROR] Import failed - check error messages above")

if __name__ == "__main__":
    main()
